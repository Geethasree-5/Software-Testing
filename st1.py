import os
import re

from openpyxl import load_workbook

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

wb = load_workbook("testdata.xlsx")
ws = wb["LoginData"]

driver = webdriver.Chrome()

driver.maximize_window()

path = "file:///" + os.path.abspath("login.html").replace("\\","/")

driver.get(path)

wait = WebDriverWait(driver,10)

email_pattern = r'^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$'

print("="*60)

for row in range(2,ws.max_row+1):

    email = ws.cell(row,1).value
    password = ws.cell(row,2).value

    print(f"\nTesting : {email}")

    # Email Validation
    if re.match(email_pattern,email):
        print("✅ Valid Email Format")
    else:
        print("❌ Invalid Email Format")

    driver.find_element(By.ID,"email").clear()
    driver.find_element(By.ID,"password").clear()

    driver.find_element(By.ID,"email").send_keys(email)
    driver.find_element(By.ID,"password").send_keys(password)

    driver.find_element(By.ID,"loginButton").click()

    msg = wait.until(
        EC.visibility_of_element_located((By.ID,"message"))
    ).text

    if email=="admin@gmail.com" and password=="admin123":

        expected="Login Successful!"

    else:

        expected="Invalid email or password!"

    if msg==expected:
        print("✅ Login Verification Passed")
    else:
        print("❌ Login Verification Failed")

print("="*60)

input("Press Enter to Exit...")

driver.quit()