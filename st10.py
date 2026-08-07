from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

from collections import defaultdict
from openpyxl import Workbook

import os

# Launch Browser
driver = webdriver.Chrome(
    service=Service(
        ChromeDriverManager().install()
    )
)

driver.maximize_window()

# Open Local HTML File
html_path = os.path.abspath(
    "employee_table1.html"
)

driver.get(
    "file:///" + html_path
)

# Read Rows
rows = driver.find_elements(
    By.XPATH,
    "//table[@id='employeeTable']/tbody/tr"
)

employees = []

employee_ids = []

duplicate_ids = set()

high_salary_employees = []

department_salary = defaultdict(list)

# Extract Data
for row in rows:

    cols = row.find_elements(
        By.TAG_NAME,
        "td"
    )

    emp_id = int(cols[0].text)
    name = cols[1].text
    department = cols[2].text
    salary = int(cols[3].text)

    employees.append([
        emp_id,
        name,
        department,
        salary
    ])

    # Duplicate Check
    if emp_id in employee_ids:
        duplicate_ids.add(emp_id)

    employee_ids.append(emp_id)

    # Salary > 75000
    if salary > 75000:

        high_salary_employees.append([
            emp_id,
            name,
            department,
            salary
        ])

    # Department Wise Salary
    department_salary[
        department
    ].append(
        salary
    )

# Department Average
department_average = {}

for dept in department_salary:

    avg = (
        sum(
            department_salary[dept]
        )
        /
        len(
            department_salary[dept]
        )
    )

    department_average[
        dept
    ] = round(avg, 2)

# Sorting Check
sorted_ids = sorted(
    employee_ids
)

is_sorted = (
    employee_ids == sorted_ids
)

# Excel Report
wb = Workbook()

# Sheet 1
sheet1 = wb.active

sheet1.title = "Employees"

sheet1.append([
    "Employee ID",
    "Name",
    "Department",
    "Salary"
])

for emp in employees:
    sheet1.append(emp)

# Sheet 2
sheet2 = wb.create_sheet(
    "Duplicate IDs"
)

sheet2.append([
    "Duplicate Employee IDs"
])

for dup in duplicate_ids:
    sheet2.append([dup])

# Sheet 3
sheet3 = wb.create_sheet(
    "Salary Above 75000"
)

sheet3.append([
    "Employee ID",
    "Name",
    "Department",
    "Salary"
])

for emp in high_salary_employees:
    sheet3.append(emp)

# Sheet 4
sheet4 = wb.create_sheet(
    "Department Average"
)

sheet4.append([
    "Department",
    "Average Salary"
])

for dept, avg in (
        department_average.items()):

    sheet4.append([
        dept,
        avg
    ])

# Sheet 5
sheet5 = wb.create_sheet(
    "Sorting Result"
)

sheet5.append([
    "Sorted By Employee ID"
])

sheet5.append([
    str(is_sorted)
])

wb.save(
    "Employee_Report.xlsx"
)

# Console Output
print("\n===== ALL EMPLOYEES =====")

for emp in employees:
    print(emp)

print("\n===== DUPLICATE IDS =====")
print(list(duplicate_ids))

print("\n===== SALARY > 75000 =====")

for emp in high_salary_employees:
    print(emp)

print("\n===== DEPARTMENT AVERAGES =====")

for dept, avg in (
        department_average.items()):

    print(
        dept,
        "->",
        avg
    )

print("\n===== SORTING CHECK =====")

if is_sorted:
    print("PASS")
else:
    print("FAIL")

print(
    "\nExcel Report Generated Successfully"
)

driver.quit()