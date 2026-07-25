import os

from openpyxl import load_workbook

from selenium import webdriver
from selenium.webdriver.common.by import By

driver=webdriver.Chrome()

driver.maximize_window()

path="file:///"+os.path.abspath("ecommerce.html").replace("\\","/")

driver.get(path)

wb=load_workbook("products.xlsx")

ws=wb["Products"]

print("="*60)

print("PRODUCT FILTER VERIFICATION")

print("="*60)

for row in range(2,ws.max_row+1):

    category=ws.cell(row,1).value
    maxprice=ws.cell(row,2).value

    print()

    print("Searching Category :",category)

    print("Maximum Price :",maxprice)

    if category=="Electronics":

        price1=int(driver.find_element(By.ID,"price1").text)

        price2=int(driver.find_element(By.ID,"price2").text)

        if price1<=maxprice:

            print("Laptop Matches Filter")

        if price2<=maxprice:

            print("Mobile Matches Filter")

    if category=="Fashion":

        price3=int(driver.find_element(By.ID,"price3").text)

        if price3<=maxprice:

            print("Shoes Matches Filter")

print()

print("="*60)

print("ADDING PRODUCTS")

print("="*60)

buttons=driver.find_elements(By.TAG_NAME,"button")

buttons[0].click()

buttons[1].click()

items=driver.find_element(By.ID,"items").text

total=driver.find_element(By.ID,"total").text

print("Items Added :",items)

print("Cart Total :",total)

expected=50000+25000

if int(total)==expected:

    print("Cart Total Verification PASSED")

else:

    print("Cart Total Verification FAILED")

print()

print("="*60)

print("Quantity Update Verification")

print("="*60)

buttons[0].click()

items=driver.find_element(By.ID,"items").text

total=driver.find_element(By.ID,"total").text

print("Updated Quantity :",items)

print("Updated Total :",total)

if int(items)==3:

    print("Quantity Updated Successfully")

else:

    print("Quantity Update Failed")

input("Press Enter to Exit...")

driver.quit()